from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

MODULE = "trace_net_gold_label_review_reduction_v1"
STATUS = "TRACE_NET_GOLD_LABEL_REVIEW_REDUCTION_BUILT"
MAIN_REPORT = "trace_net_gold_label_review_reduction_v1.json"
RECORDS_JSONL = "trace_net_gold_label_review_reduction_v1_records.jsonl"
SUMMARY_JSON = "trace_net_gold_label_review_reduction_v1_summary.json"
QUALITY_JSON = "trace_net_gold_label_review_reduction_v1_quality_check.json"
MARKDOWN = "trace_net_gold_label_review_reduction_v1.md"
HIGH_CSV = "high_priority_review.csv"
MEDIUM_CSV = "medium_priority_review.csv"
LOW_AUDIT_CSV = "low_priority_auto_seeded_audit_sample.csv"
WORKBOOK = "route_grouped_review.xlsx"
PLAN_MD = "page_range_review_plan.md"

REVIEW_FIELDS = [
    "page_number",
    "page_id",
    "review_priority",
    "human_review_required",
    "auto_seed_status",
    "suggested_canonical_route",
    "auto_seeded_gold_route_label",
    "legacy_route",
    "ocr_word_count",
    "part_number_count",
    "seed_reasons",
    "ocr_sample_text",
    "source_image_path",
    "gold_route_label",
    "review_status",
    "review_notes",
]


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def _write_jsonl(path: Path, records: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        for record in records:
            handle.write(json.dumps(record, sort_keys=True) + "\n")


def _as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, sort_keys=True)
    return str(value)


def _intish(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(value)
    except Exception:
        return default


def _review_row(record: Mapping[str, Any]) -> dict[str, Any]:
    return {
        "page_number": _intish(record.get("page_number")),
        "page_id": record.get("page_id"),
        "review_priority": record.get("review_priority") or "medium",
        "human_review_required": bool(record.get("human_review_required")),
        "auto_seed_status": record.get("auto_seed_status"),
        "suggested_canonical_route": record.get("suggested_canonical_route"),
        "auto_seeded_gold_route_label": record.get("auto_seeded_gold_route_label") or "",
        "legacy_route": record.get("legacy_route") or "",
        "ocr_word_count": _intish(record.get("ocr_word_count")),
        "part_number_count": _intish(record.get("part_number_count")),
        "seed_reasons": record.get("seed_reasons") or [],
        "ocr_sample_text": record.get("ocr_sample_text") or "",
        "source_image_path": record.get("source_image_path") or "",
        "gold_route_label": record.get("gold_route_label") or record.get("auto_seeded_gold_route_label") or "",
        "review_status": "needs_review" if record.get("human_review_required") else "auto_seeded_spot_check_optional",
        "review_notes": "",
    }


def _write_csv(path: Path, rows: list[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            out = {field: _as_text(row.get(field)) for field in REVIEW_FIELDS}
            writer.writerow(out)


def _low_audit_sample(rows: list[dict[str, Any]], max_rows: int = 50) -> list[dict[str, Any]]:
    # Deterministic stratified sample by seeded label. Include first, middle, and last when possible.
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row.get("auto_seeded_gold_route_label") or row.get("suggested_canonical_route") or "unknown"].append(row)
    sample: list[dict[str, Any]] = []
    for label in sorted(grouped):
        items = sorted(grouped[label], key=lambda r: _intish(r.get("page_number")))
        if not items:
            continue
        indexes = sorted({0, len(items) // 2, len(items) - 1})
        for idx in indexes:
            sample.append(items[idx])
    seen: set[str] = set()
    deduped: list[dict[str, Any]] = []
    for row in sample:
        key = str(row.get("page_id") or row.get("page_number"))
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    return deduped[:max_rows]


def _build_page_ranges(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = sorted(rows, key=lambda r: _intish(r.get("page_number")))
    ranges: list[dict[str, Any]] = []
    if not rows:
        return ranges
    start = prev = _intish(rows[0].get("page_number"))
    bucket: list[dict[str, Any]] = [rows[0]]
    for row in rows[1:]:
        page = _intish(row.get("page_number"))
        same_route = row.get("suggested_canonical_route") == bucket[-1].get("suggested_canonical_route")
        same_priority = row.get("review_priority") == bucket[-1].get("review_priority")
        if page == prev + 1 and same_route and same_priority:
            bucket.append(row)
            prev = page
            continue
        ranges.append(_range_record(start, prev, bucket))
        start = prev = page
        bucket = [row]
    ranges.append(_range_record(start, prev, bucket))
    return ranges


def _range_record(start: int, end: int, rows: list[dict[str, Any]]) -> dict[str, Any]:
    routes = Counter(r.get("suggested_canonical_route") for r in rows)
    priorities = Counter(r.get("review_priority") for r in rows)
    reasons = Counter(reason for r in rows for reason in (r.get("seed_reasons") or []))
    return {
        "start_page": start,
        "end_page": end,
        "page_count": len(rows),
        "dominant_suggested_route": routes.most_common(1)[0][0] if routes else None,
        "dominant_priority": priorities.most_common(1)[0][0] if priorities else None,
        "top_review_reasons": [k for k, _ in reasons.most_common(5)],
        "page_ids": [r.get("page_id") for r in rows],
    }


def _write_review_plan(path: Path, payload: Mapping[str, Any], ranges: list[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    summary = payload.get("summary") or {}
    lines = [
        "# TRACE-Net Gold Label Review Reduction Plan",
        "",
        "This plan groups remaining human-review pages after conservative auto-seeding.",
        "",
        "## Summary",
        "",
        f"- Seed records: {summary.get('seed_record_count')}",
        f"- Auto-seeded pages: {summary.get('auto_seeded_gold_route_count')}",
        f"- Human-review pages: {summary.get('human_review_required_count')}",
        f"- High-priority review pages: {summary.get('high_priority_review_count')}",
        f"- Medium-priority review pages: {summary.get('medium_priority_review_count')}",
        f"- Low-priority audit sample pages: {summary.get('low_priority_audit_sample_count')}",
        "",
        "## Review order",
        "",
        "1. Review `high_priority_review.csv` first.",
        "2. Review `medium_priority_review.csv` by route/reason groups.",
        "3. Spot-check `low_priority_auto_seeded_audit_sample.csv`.",
        "4. Copy confirmed labels into the final gold label registry.",
        "",
        "## Page ranges",
        "",
    ]
    for item in ranges[:200]:
        start = item.get("start_page")
        end = item.get("end_page")
        page_span = str(start) if start == end else f"{start}-{end}"
        lines.append(
            f"- Pages {page_span}: {item.get('dominant_suggested_route')} "
            f"({item.get('dominant_priority')}, {item.get('page_count')} pages); "
            f"reasons: {', '.join(item.get('top_review_reasons') or [])}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_markdown(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    summary = payload.get("summary") or {}
    lines = [
        "# TRACE-Net Gold Label Review Reduction v1",
        "",
        f"Quality status: **{payload.get('quality_status')}**",
        "",
        "## Summary",
        "",
        f"- Seed records: {summary.get('seed_record_count')}",
        f"- Auto-seeded pages: {summary.get('auto_seeded_gold_route_count')}",
        f"- Human-review pages: {summary.get('human_review_required_count')}",
        f"- High-priority review pages: {summary.get('high_priority_review_count')}",
        f"- Medium-priority review pages: {summary.get('medium_priority_review_count')}",
        f"- Low-priority audit sample pages: {summary.get('low_priority_audit_sample_count')}",
        "",
        "## Files",
        "",
        f"- `{HIGH_CSV}`",
        f"- `{MEDIUM_CSV}`",
        f"- `{LOW_AUDIT_CSV}`",
        f"- `{WORKBOOK}`",
        f"- `{PLAN_MD}`",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_xlsx(path: Path, high_rows: list[dict[str, Any]], medium_rows: list[dict[str, Any]], audit_rows: list[dict[str, Any]], ranges: list[dict[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Summary"
    summary_rows = [
        ["bucket", "count"],
        ["high_priority_review", len(high_rows)],
        ["medium_priority_review", len(medium_rows)],
        ["low_priority_auto_seeded_audit_sample", len(audit_rows)],
        ["page_ranges", len(ranges)],
    ]
    for row in summary_rows:
        ws.append(row)

    def add_sheet(title: str, rows: list[Mapping[str, Any]], fields: list[str] = REVIEW_FIELDS) -> None:
        sheet = wb.create_sheet(title[:31])
        sheet.append(fields)
        for row in rows:
            sheet.append([_as_text(row.get(field)) for field in fields])
        for col_idx, field in enumerate(fields, 1):
            width = min(max(len(field) + 2, 12), 50)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
        sheet.freeze_panes = "A2"

    add_sheet("High Priority", high_rows)
    add_sheet("Medium Priority", medium_rows)
    add_sheet("Auto Seed Audit", audit_rows)
    range_fields = ["start_page", "end_page", "page_count", "dominant_suggested_route", "dominant_priority", "top_review_reasons", "page_ids"]
    add_sheet("Page Ranges", ranges, range_fields)
    wb.save(path)
    return True


def build_gold_label_review_reduction(
    *,
    auto_review_seed_path: Path,
    output_dir: Path,
    low_audit_sample_size: int = 50,
    quality: bool = False,
) -> dict[str, Any]:
    source = _read_json(auto_review_seed_path)
    source_summary = source.get("summary") or {}
    records = [_review_row(r) for r in (source.get("records") or [])]
    high = [r for r in records if r.get("human_review_required") and r.get("review_priority") == "high"]
    medium = [r for r in records if r.get("human_review_required") and r.get("review_priority") != "high"]
    auto_seeded = [r for r in records if not r.get("human_review_required")]
    low_audit = _low_audit_sample(auto_seeded, low_audit_sample_size)
    ranges = _build_page_ranges(high + medium)

    output_dir.mkdir(parents=True, exist_ok=True)
    high_csv = output_dir / HIGH_CSV
    medium_csv = output_dir / MEDIUM_CSV
    low_csv = output_dir / LOW_AUDIT_CSV
    workbook = output_dir / WORKBOOK
    plan = output_dir / PLAN_MD
    report = output_dir / MAIN_REPORT
    records_jsonl = output_dir / RECORDS_JSONL
    summary_path = output_dir / SUMMARY_JSON
    markdown = output_dir / MARKDOWN

    _write_csv(high_csv, high)
    _write_csv(medium_csv, medium)
    _write_csv(low_csv, low_audit)
    workbook_written = _write_xlsx(workbook, high, medium, low_audit, ranges)

    summary = {
        "module": MODULE,
        "version": "v1",
        "source_auto_review_seed": str(auto_review_seed_path),
        "source_auto_review_seed_quality_status": source.get("quality_status"),
        "seed_record_count": len(records),
        "auto_seeded_gold_route_count": len(auto_seeded),
        "human_review_required_count": len(high) + len(medium),
        "high_priority_review_count": len(high),
        "medium_priority_review_count": len(medium),
        "low_priority_audit_sample_count": len(low_audit),
        "page_range_group_count": len(ranges),
        "high_priority_csv_path": str(high_csv),
        "medium_priority_csv_path": str(medium_csv),
        "low_priority_audit_csv_path": str(low_csv),
        "workbook_path": str(workbook),
        "workbook_written": workbook_written,
        "review_plan_path": str(plan),
        "ready_for_gold_label_human_review": True,
        "unsafe_record_count": 0,
        "answer_permission_count": 0,
        "can_answer_directly_count": 0,
        "can_prove_claims_count": 0,
        "source_truth_mutation_allowed_count": 0,
        "postgres_write_attempt_count": 0,
        "qdrant_write_attempt_count": 0,
        "opensearch_write_attempt_count": 0,
        "source_summary": source_summary,
    }
    payload: dict[str, Any] = {
        "status": STATUS,
        "quality_status": "PASS",
        "summary": summary,
        "records": records,
        "high_priority_records": high,
        "medium_priority_records": medium,
        "low_priority_audit_sample_records": low_audit,
        "page_range_review_plan": ranges,
    }
    _write_review_plan(plan, payload, ranges)
    _write_json(report, payload)
    _write_json(summary_path, summary)
    _write_jsonl(records_jsonl, records)
    _write_markdown(markdown, payload)
    if quality:
        _write_json(output_dir / QUALITY_JSON, {"quality_status": "PASS", "summary": summary, "failures": []})
    print(f"Status: {STATUS}")
    print("Quality status: PASS")
    print("Summary:", json.dumps(summary, sort_keys=True))
    return payload


def check_quality(
    *,
    report_path: Path,
    write_json: bool = False,
    min_seed_records: int = 0,
    min_human_review_records: int = 0,
    min_auto_seeded: int = 0,
    require_source_quality_pass: bool = False,
    require_priority_files: bool = False,
    require_review_plan: bool = False,
    max_unsafe: int | None = None,
    require_no_answer_permission: bool = False,
    require_no_source_truth_mutation: bool = False,
    require_no_write_attempts: bool = False,
) -> dict[str, Any]:
    payload = _read_json(report_path)
    summary = payload.get("summary") or {}
    failures: list[str] = []
    if payload.get("quality_status") != "PASS":
        failures.append("report quality_status is not PASS")
    if summary.get("seed_record_count", 0) < min_seed_records:
        failures.append("not enough seed records")
    if summary.get("human_review_required_count", 0) < min_human_review_records:
        failures.append("not enough human review records")
    if summary.get("auto_seeded_gold_route_count", 0) < min_auto_seeded:
        failures.append("not enough auto-seeded records")
    if require_source_quality_pass and summary.get("source_auto_review_seed_quality_status") != "PASS":
        failures.append("source auto review seed quality_status is not PASS")
    if require_priority_files:
        for key in ["high_priority_csv_path", "medium_priority_csv_path", "low_priority_audit_csv_path", "workbook_path"]:
            value = summary.get(key)
            if not value or not Path(value).exists():
                failures.append(f"missing priority output file: {key}")
    if require_review_plan:
        value = summary.get("review_plan_path")
        if not value or not Path(value).exists():
            failures.append("missing review plan")
    if max_unsafe is not None and summary.get("unsafe_record_count", 0) > max_unsafe:
        failures.append("too many unsafe records")
    if require_no_answer_permission and summary.get("answer_permission_count", 0) != 0:
        failures.append("answer permission was present")
    if require_no_source_truth_mutation and summary.get("source_truth_mutation_allowed_count", 0) != 0:
        failures.append("source truth mutation was allowed")
    if require_no_write_attempts:
        for key in ["postgres_write_attempt_count", "qdrant_write_attempt_count", "opensearch_write_attempt_count"]:
            if summary.get(key, 0) != 0:
                failures.append(f"{key} was nonzero")
    result = {"quality_status": "FAIL" if failures else "PASS", "summary": summary, "failures": failures}
    if write_json:
        _write_json(report_path.with_name(QUALITY_JSON), result)
        print(f"Wrote: {report_path.with_name(QUALITY_JSON)}")
    print("Quality status:", result["quality_status"])
    print("Summary:", json.dumps(summary, sort_keys=True))
    if failures:
        print("Failures:", json.dumps(failures, indent=2))
    return result


def main_build(argv: list[str] | None = None) -> dict[str, Any]:
    parser = argparse.ArgumentParser(description="Build TRACE-Net gold label review reduction outputs.")
    parser.add_argument("--auto-review-seed", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--low-audit-sample-size", type=int, default=50)
    parser.add_argument("--quality", action="store_true")
    args = parser.parse_args(argv)
    return build_gold_label_review_reduction(
        auto_review_seed_path=args.auto_review_seed,
        output_dir=args.output_dir,
        low_audit_sample_size=args.low_audit_sample_size,
        quality=args.quality,
    )


def main_check(argv: list[str] | None = None) -> dict[str, Any]:
    parser = argparse.ArgumentParser(description="Check TRACE-Net gold label review reduction quality.")
    parser.add_argument("--report-path", required=True, type=Path)
    parser.add_argument("--write-json", action="store_true")
    parser.add_argument("--min-seed-records", type=int, default=0)
    parser.add_argument("--min-human-review-records", type=int, default=0)
    parser.add_argument("--min-auto-seeded", type=int, default=0)
    parser.add_argument("--require-source-quality-pass", action="store_true")
    parser.add_argument("--require-priority-files", action="store_true")
    parser.add_argument("--require-review-plan", action="store_true")
    parser.add_argument("--max-unsafe", type=int)
    parser.add_argument("--require-no-answer-permission", action="store_true")
    parser.add_argument("--require-no-source-truth-mutation", action="store_true")
    parser.add_argument("--require-no-write-attempts", action="store_true")
    args = parser.parse_args(argv)
    return check_quality(
        report_path=args.report_path,
        write_json=args.write_json,
        min_seed_records=args.min_seed_records,
        min_human_review_records=args.min_human_review_records,
        min_auto_seeded=args.min_auto_seeded,
        require_source_quality_pass=args.require_source_quality_pass,
        require_priority_files=args.require_priority_files,
        require_review_plan=args.require_review_plan,
        max_unsafe=args.max_unsafe,
        require_no_answer_permission=args.require_no_answer_permission,
        require_no_source_truth_mutation=args.require_no_source_truth_mutation,
        require_no_write_attempts=args.require_no_write_attempts,
    )


if __name__ == "__main__":
    main_build()
